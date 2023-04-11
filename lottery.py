from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font


def check_input_validation():
    while True:
        custom_response = input("Enter your desired draw number: (1 / 2): ")
        try:
            custom_response = int(custom_response)
            if custom_response == 1 or custom_response == 2:
                return custom_response
            else:
                print("Enter a valid number \"1\" or \"2\"!!")
                continue
        except ValueError:
            print("Enter a number")


def read_from_file(draw_number):

    if draw_number == 1:
        draw_number = 3
    else:
        draw_number = 4

    wb = load_workbook("toto_statistic.xlsx")
    sheets = wb.sheetnames
    sheet_one = wb[sheets[0]]

    draw_numbers = []

    for col in sheet_one.iter_cols(min_col=draw_number, max_col=draw_number, min_row=4):
        for cell in col:
            cur_values = cell.value
            values = tuple(int(i) for i in cur_values.split(","))
            draw_numbers.append(values)

    return draw_numbers


def check_for_lucky_numbers(cur_draw, next_draw, numb_to_iterate):
    not_found_numbers = []
    lucky_numbers = []
    edited_draw = next_draw[:numb_to_iterate]

    for j in range(numb_to_iterate):
        if cur_draw[j] not in edited_draw:
            not_found_numbers.append(True)
        else:
            lucky_numbers.append(cur_draw[j])

    return not_found_numbers, lucky_numbers


def add_to_pattern(pattern_option, numbers, cur_draw):

    if numbers not in pattern_option:
        pattern_option[numbers] = []
    pattern_option[numbers].append(cur_draw)

    return pattern_option


def iterate_for_rest_numbers(numb_to_iterate, cur_draw, next_draw):
    other_lucky_numbs = []
    edited_draw = next_draw[numb_to_iterate:len(next_draw)]

    for i in range(numb_to_iterate, len(cur_draw)):
        if cur_draw[i] in edited_draw:
            other_lucky_numbs.append(cur_draw[i])

    return other_lucky_numbs


def write_data_to_file(file_name, pattern_of_numbers):
    wb = load_workbook(file_name)

    sheets = wb.sheetnames
    sheet_one = wb[sheets[0]]

    count_rows = 2
    for key, val in sorted(pattern_of_numbers.items(), key=lambda x: (-len(x[1]), sorted(x[0]))):
        for row in sheet_one.iter_rows(min_row=count_rows, max_row=count_rows):
            sheet_one[row[0].coordinate] = f"{', '.join([str(char) for char in key])}"
            sheet_one[row[1].coordinate] = f"{sorted(val)}"
            sheet_one[row[2].coordinate] = f"{len(val)}"

        count_rows += 1
    wb.save(file_name)


def get_highest_winning_combinations(pattern_of_numbers, numb_of_pattern):
    file_to_open = create_file(numb_of_pattern)
    write_data_to_file(file_to_open, pattern_of_numbers)


def create_file(number):
    # To implement - align of the text, auto width, text for the saved file,
    # tab name auto, check if the file exists -> if yes to ask to rewrite or make another one
    # less code for the main stats

    work_book = Workbook()
    main_sheet = work_book.active
    main_sheet.title = "2018"

    main_sheet["A1"] = "Numbers"
    main_sheet["B1"] = "Draw Order"
    main_sheet["C1"] = "Number of hits"

    main_sheet.column_dimensions["A"].width = 20
    main_sheet.column_dimensions["B"].width = 20
    main_sheet.column_dimensions["C"].width = 25
    font = Font(color="00FF0000", size="17", bold=True)
    a_row = main_sheet["A1"]
    a_row.font = font
    b_row = main_sheet["B1"]
    b_row.font = font
    c_row = main_sheet["C1"]
    c_row.font = font

    name_of_file = None

    if number == 0:
        name_of_file = "Pattern_Three"
    elif number == 1:
        name_of_file = "Pattern_Four"
    elif number == 2:
        name_of_file = "Pattern_Five"

    work_book.save(f"{name_of_file}.xlsx")

    return f"{name_of_file}.xlsx"


def check_main_patterns(patterns):
    pattern_of_three = {}
    pattern_of_four = {}
    pattern_of_five = {}

    for i in range(len(patterns)):

        # iterate only for first 3 numbers, so we save time
        number_to_iterate = int(len(patterns[i]) / 2) + 1
        current_draw = patterns[i]

        number_of_draws_to_check = (len(patterns) - i) - 1

        for j in range(number_of_draws_to_check):
            next_pattern_number = i + (j + 1)
            next_draw = patterns[next_pattern_number]

            not_found_numbers, lucky_numbers = check_for_lucky_numbers(current_draw, next_draw, number_to_iterate)

            if all(not_found_numbers) and len(not_found_numbers) == number_to_iterate:
                continue
            else:
                result = iterate_for_rest_numbers(number_to_iterate, current_draw, next_draw)

                lucky_numbers.extend(result)

                numbers = tuple(sorted(lucky_numbers))

                if len(lucky_numbers) == 5:
                    pattern_of_five = add_to_pattern(pattern_of_five, numbers, next_pattern_number)
                elif len(lucky_numbers) == 4:
                    pattern_of_four = add_to_pattern(pattern_of_four, numbers, next_pattern_number)
                elif len(lucky_numbers) == 3:
                    pattern_of_three = add_to_pattern(pattern_of_three, numbers, next_pattern_number)

    objects = [pattern_of_three, pattern_of_four, pattern_of_five]
    for pat in (range(len(objects))):
        get_highest_winning_combinations(objects[pat], pat)


def main():

    numb = check_input_validation()

    patterns = read_from_file(numb)

    check_main_patterns(patterns)

    print("Done!")


if __name__ == "__main__":
    main()
